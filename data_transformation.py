import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import tkinter
from tkinter import filedialog
from datetime import date


def data_transformation(rutas, meses):

    ruta_excel_base1 = rutas[0]
    ruta_excel_base2 = rutas[1]
    ruta_excel_base3 = rutas[2]

    mes_medicion = meses[0]
    mes_anterior = meses[1]
    mes_antepasado = meses[2]

    dataframe1 = pd.read_excel(
        ruta_excel_base1, sheet_name="Priorizacion " + mes_medicion
    )
    dataframe1[mes_antepasado] = 1

    dataframe2 = pd.read_excel(
        ruta_excel_base2, sheet_name="Priorizacion " + mes_anterior
    )
    dataframe2[mes_anterior] = 1

    dataframe3 = pd.read_excel(
        ruta_excel_base3, sheet_name="Priorizacion " + mes_antepasado
    )
    dataframe3[mes_medicion] = 1

    dataframe_fusionado = pd.merge(dataframe1, dataframe2, on="NODO", how="outer")
    dataframe_fusionado = pd.merge(
        dataframe_fusionado, dataframe3, on="NODO", how="outer"
    )

    dataframe_fusionado[mes_antepasado].fillna(0, inplace=True)
    dataframe_fusionado[mes_anterior].fillna(0, inplace=True)
    dataframe_fusionado[mes_medicion].fillna(0, inplace=True)

    columnas = ["NODO", mes_antepasado, mes_anterior, mes_medicion]

    dataframe_fusionado = dataframe_fusionado[columnas]

    # Define la función para evaluar cada fila y asignar un texto y un valor numérico
    def evaluar_fila(fila):
        if (
            fila[mes_medicion] == 1
            and fila[mes_anterior] == 1
            and fila[mes_antepasado] == 1
        ):
            return ("Recurrente 3 meses", 3)
        elif fila[mes_medicion] == 1 and (
            fila[mes_anterior] == 1 or fila[mes_antepasado] == 1
        ):
            return ("Mes de medicion si + 1 (cualquiera)", 4)
        elif (
            fila[mes_medicion] == 0
            and fila[mes_anterior] == 1
            and fila[mes_antepasado] == 1
        ):
            return ("Mes de medicion no + 2", 5)
        elif (
            fila[mes_medicion] == 1
            and fila[mes_anterior] == 0
            and fila[mes_antepasado] == 0
        ):
            return ("Mes de medicion si + 0", 6)
        elif fila[mes_medicion] == 0 and (
            fila[mes_anterior] == 1 or fila[mes_antepasado] == 1
        ):
            return ("Mes de medicion no + 1 (cualquiera)", 7)

    # Aplica la función a cada fila del DataFrame utilizando apply()
    dataframe_fusionado[["Caso", "Priorizacion"]] = dataframe_fusionado.apply(
        lambda fila: pd.Series(evaluar_fila(fila), index=["Caso", "Priorizacion"]),
        axis=1,
    )

    dataframe_fusionado_ordenado = dataframe_fusionado.sort_values(by="Priorizacion")

    dataframe_fusionado_ordenado.to_excel("file.xlsx", index=False)

    libro = openpyxl.load_workbook("file.xlsx")
    hoja = libro.active
    hoja.title = "Prio Fijo"

    # Obtén la fila específica
    fila_especifica = hoja[1062]

    # Define el estilo de resaltado (por ejemplo, fondo amarillo)
    estilo_resaltado = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    # Aplica el estilo a cada celda de la fila
    for celda in fila_especifica:
        celda.fill = estilo_resaltado

    # Crear una ventana modal
    window = tkinter.Tk()
    window.withdraw()
    window.attributes("-topmost", True)

    # Agregar el widget para seleccionar el archivo
    ruta_archivo = filedialog.asksaveasfilename(
        title="Guardar archivo",
        filetypes=[("Archivos de Excel", "*.xlsx")],
        defaultextension=".xlsx",
        initialfile="BackLog_Fijo_{}".format(date.today()),
    )

    # Si el usuario seleccionó un archivo
    if ruta_archivo:
        # Guardar el libro de Excel
        libro.save(ruta_archivo)

    # Cerrar la ventana modal
    window.destroy()
