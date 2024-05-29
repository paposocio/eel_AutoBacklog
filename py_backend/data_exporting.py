import openpyxl, pandas
from openpyxl.styles import PatternFill
import tkinter
from tkinter import filedialog
from datetime import date
from configparser import ConfigParser


def data_export(option):

    if option == 1:
        libro = openpyxl.load_workbook("file_preview.xlsx")
        hoja = libro.active
        hoja.title = "Prio Fijo"

        config = ConfigParser()
        config.read("settings.ini", encoding="utf-8")

        # Obtén la fila específica
        fila_especifica = hoja[(int(config["ConfigFijo"]["topenodos"]) + 2)]

        # Define el estilo de resaltado (por ejemplo, fondo amarillo)
        estilo_resaltado = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        # Aplica el estilo a cada celda de la fila
        for celda in fila_especifica:
            celda.fill = estilo_resaltado

        # Crear una ventana modal
        window = tkinter.Tk()
        window.withdraw()
        window.attributes("-topmost", True)

        # Agregar el widget para seleccionar el archivo
        ruta_archivo = filedialog.asksaveasfilename(
            title="Guardar archivo",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            defaultextension=".xlsx",
            initialfile="BackLog_Fijo_{}".format(date.today()),
        )

        # Si el usuario seleccionó un archivo
        if ruta_archivo:
            # Guardar el libro de Excel
            libro.save(ruta_archivo)

        # Cerrar la ventana modal
        window.destroy()

    elif option == 2:
        # Crear una ventana modal
        window = tkinter.Tk()
        window.withdraw()
        window.attributes("-topmost", True)

        # Agregar el widget para seleccionar el archivo
        ruta_archivo = filedialog.asksaveasfilename(
            title="Guardar archivo",
            filetypes=[("Archivo CSV", "*.csv")],
            defaultextension=".csv",
            initialfile="BackLog_Fijo_{}".format(date.today()),
        )

        # Si el usuario seleccionó un archivo
        if ruta_archivo:
            dataframe = pandas.read_excel("file_preview.xlsx")
            dataframe.to_csv(ruta_archivo, index=False, header=True)
        else:
            return "Sin archivo seleccionado"

        # Cerrar la ventana modal
        window.destroy()

    elif option == 3:
        dataframe = pandas.read_excel("file_preview.xlsx")
        dataframe.to_clipboard(index=False)


def data_export_mobile(option):

    if option == 1:
        libro = openpyxl.load_workbook("file_preview_mobile.xlsx")
        hoja = libro.active
        hoja.title = "Prio Movil"

        config = ConfigParser()
        config.read("settings.ini", encoding="utf-8")

        # Obtén la fila específica
        fila_especifica = hoja[(int(config["ConfigMovil"]["topestacionbase"]) + 2)]

        # Define el estilo de resaltado (por ejemplo, fondo amarillo)
        estilo_resaltado = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        # Aplica el estilo a cada celda de la fila
        for celda in fila_especifica:
            celda.fill = estilo_resaltado

        # Crear una ventana modal
        window = tkinter.Tk()
        window.withdraw()
        window.attributes("-topmost", True)

        # Agregar el widget para seleccionar el archivo
        ruta_archivo = filedialog.asksaveasfilename(
            title="Guardar archivo",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            defaultextension=".xlsx",
            initialfile="BackLog_Movil_{}".format(date.today()),
        )

        # Si el usuario seleccionó un archivo
        if ruta_archivo:
            # Guardar el libro de Excel
            libro.save(ruta_archivo)

        # Cerrar la ventana modal
        window.destroy()

    elif option == 2:
        # Crear una ventana modal
        window = tkinter.Tk()
        window.withdraw()
        window.attributes("-topmost", True)

        # Agregar el widget para seleccionar el archivo
        ruta_archivo = filedialog.asksaveasfilename(
            title="Guardar archivo",
            filetypes=[("Archivo CSV", "*.csv")],
            defaultextension=".csv",
            initialfile="BackLog_Movil_{}".format(date.today()),
        )

        # Si el usuario seleccionó un archivo
        if ruta_archivo:
            dataframe = pandas.read_excel("file_preview_mobile.xlsx")
            dataframe.to_csv(ruta_archivo, index=False, header=True)
        else:
            return "Sin archivo seleccionado"

        # Cerrar la ventana modal
        window.destroy()

    elif option == 3:
        dataframe = pandas.read_excel("file_preview_mobile.xlsx")
        dataframe.to_clipboard(index=False)